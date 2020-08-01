# importing openpyxl module 
import openpyxl 

# To create a new workbook to store results
wb = openpyxl.Workbook()
sheet = wb.active
# Column Headers 
c1 = sheet['A1']
c1.value = "Event_ID"
c2 = sheet['B1']
c2.value = "Facility"
c3 = sheet['C1']
c3.value = "Dir"
c4 = sheet['D1']
c4.value = "Nearest INRIX Seg. ID"
c5 = sheet['E1']
c5.value = "Nearest Wavetronix ID"

# Give the name & location of the data source files
path1 = "C:\\Users\\lenovo\\Desktop\\Project_temp\\codes\\sensors.xlsx"
path2 = "C:\\Users\\lenovo\\Desktop\\Project_temp\\codes\\inrix.xlsx"
path3 = "C:\\Users\\lenovo\\Desktop\\Project_temp\\codes\\incidents.xlsx"

# workbook object is created 
wb_sensors = openpyxl.load_workbook(path1)
wb_inrix = openpyxl.load_workbook(path2)
wb_incidents = openpyxl.load_workbook(path3)

sheet_sensors = wb_sensors.active
sheet_inrix = wb_inrix.active
sheet_incidents = wb_incidents.active

# Main Code
for i in range(2, sheet_incidents.max_row + 1):
    event_id = sheet_incidents.cell(row = i, column = 2)
    facility = sheet_incidents.cell(row = i, column = 3)
    direction = sheet_incidents.cell(row = i, column = 4)
    lati = sheet_incidents.cell(row = i, column = 8)
    loni = sheet_incidents.cell(row = i, column = 9)

    result1 = sheet.cell(row = i, column = 1)
    result1.value = event_id.value
    result2 = sheet.cell(row = i, column = 2)
    result2.value = facility.value
    result3 = sheet.cell(row = i, column = 3)
    result3.value = direction.value
    result4 = sheet.cell(row = i, column = 4)
    result4.value = "NA"
    result5 = sheet.cell(row = i, column = 5)
    result5.value = "NA"

    # Code to get Nearest Inrix Seg. ID
    for j in range(2, sheet_inrix.max_row + 1):
        inrix_id = sheet_inrix.cell(row = j, column = 1)
        inrix_lat = sheet_inrix.cell(row = j, column = 2)
        inrix_lon = sheet_inrix.cell(row = j, column = 3)
        inrix_fac = sheet_inrix.cell(row = j, column = 4)
        inrix_dir = sheet_inrix.cell(row = j, column = 5)

        if facility.value == inrix_fac.value and direction.value == inrix_dir.value :
            if abs(lati.value - inrix_lat.value) <= 0.016 and abs(loni.value - inrix_lon.value) <= 0.016 :
                result4.value = inrix_id.value
                break

    # Code to get Nearest Wavetronix ID
    for k in range(2, sheet_sensors.max_row + 1):
        wave_id = sheet_sensors.cell(row = k, column = 1)
        wave_lat = sheet_sensors.cell(row = k, column = 2)
        wave_lon = sheet_sensors.cell(row = k, column = 3)
        wave_fac = sheet_sensors.cell(row = k, column = 4)
        wave_dir = sheet_sensors.cell(row = k, column = 5)

        if facility.value == wave_fac.value and direction.value == wave_dir.value :
            if abs(lati.value - wave_lat.value) <= 0.016 and abs(loni.value - wave_lon.value) <= 0.016 :
                result5.value = wave_id.value
                break
            
# End of Main Code

# Give name & location of result file
wb.save("C:\\Users\\lenovo\\Desktop\\Project_temp\\codes\\results\\result.xlsx")